#!/usr/bin/env python
# coding: utf-8

# In[2]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from datetime import datetime, timedelta
import os
import openpyxl
import editpyxl
import excel2img
import xlwt
from datetime import datetime
import time


# In[2]:
def actual(INPATH,OUTPATH):

    wb1=pd.read_excel(INPATH+'2G_KPI_Report_Hourly_IBS.xlsx',sheet_name="Report 3")
    wb2=pd.read_excel(INPATH+"2G_Relation-wise_HOSR_Hourly_IBS.xlsx",sheet_name="Report 1")
    wb3=pd.read_excel(INPATH+'3G_KPI_Report_Hourly_IBS.xlsx',sheet_name="3G Cellwise Hourly")
    wb4=pd.read_excel(INPATH+"3G_Relation-wise_HOSR_Hourly_IBS.xlsx",sheet_name="Report 1")
    wb5=pd.read_excel(INPATH+'4G_KPI_Report_Hourly_IBS.xlsx',sheet_name="4G KPI Table")
    print("Reading of 5 input excel sheet completed")
    wb6=pd.read_excel(INPATH+"4G_Relation-wise_HOSR_Hourly_IBS.xlsx",sheet_name="HOSR")
    wb7=pd.read_excel(INPATH+"4G_UL_RSSI_Houly_IBS.xlsx",sheet_name="UL RSSI")

    wb8=pd.read_excel(INPATH+'CA_Performance_Cellwise_Hourly.xlsx',sheet_name="Peak Throughput")

    wb9=pd.read_excel(INPATH+'3G_Throughput_Hourly.xlsx',sheet_name="3G KPI Table PLMN Level")

    wb10=pd.read_excel(INPATH+'CA_Performance_Cellwise_Hourly.xlsx',sheet_name="CA")
    print("Reading of 10 input excel sheet completed")

    wb11g=pd.read_excel(INPATH+'2G_O_D_to_I_D_Relation-wise_HOSR_Daily_IBS.xlsx',sheet_name="Report 1")

    wb12=pd.read_excel(INPATH+'4G_O_D_to_I_D_Relation-wise_HOSR_Daily_IBS.xlsx',sheet_name="HOSR")


    wb11=pd.read_excel(INPATH+'3G_Relation-wise_HOSR_Hourly_IBS-Analysis-1.xlsx',sheet_name="Report 1")
    wb21=pd.read_excel(INPATH+'3G_Relation-wise_HOSR_Hourly_IBS-Analysis-2.xlsx',sheet_name="Report 1")
    wb31=pd.read_excel(INPATH+'3G_Relation-wise_HOSR_Hourly_IBS-Analysis-3.xlsx',sheet_name="Report 1")
    wb41=pd.read_excel(INPATH+'3G_Relation-wise_HOSR_Hourly_IBS-Analysis-4.xlsx',sheet_name="Report 1")
    print("Reading of 16 input excel sheet completed")
#SRS=pd.read_excel("SRS_DATABASE.xlsx",sheet_name="Outdoor Site Data")


    wb1=wb1.replace("#DIV/0","")
    wb2=wb2.replace("#DIV/0","")
    wb3=wb3.replace("#DIV/0","")
    wb4=wb4.replace("#DIV/0","")
    wb5=wb5.replace("#DIV/0","")
    wb6=wb6.replace("#DIV/0","")
    wb7=wb7.replace("#DIV/0","")
    wb8=wb8.replace("#DIV/0",0)
    wb9=wb9.replace("#DIV/0",0)
#wb8=wb8.fillna(0)
#wb9=wb9.fillna(0)
    wb10=wb10.replace("#DIV/0","")
    wb11g=wb11g.replace("#DIV/0","")
    wb12=wb12.replace("#DIV/0","")

    wb1.columns=wb1.iloc[0]
    #wb1.columns
    wb1=wb1.reindex(wb1.index.drop(0)).reset_index(drop=True)
    GSM=wb1.iloc[:,1:]

#wb2_2=wb2
    wb2a=wb2.iloc[2:,1:]
#wb2a
    wb2a.columns=wb2a.iloc[0]
#wb2a.columns
    GSM_HO=wb2a.reindex(wb2a.index.drop(2)).reset_index(drop=True)

    wb3.columns=wb3.iloc[0]
#wb3.columns
    wb3=wb3.reindex(wb3.index.drop(0)).reset_index(drop=True)
    WCDMA=wb3.iloc[:,1:]

#wb4_4=wb4
    wb4a=wb4.iloc[2:,1:]
    wb4a.columns=wb4a.iloc[0]
    #wb4a.columns
    WCDMA_HO=wb4a.reindex(wb4a.index.drop(2)).reset_index(drop=True)

    wb5.columns=wb5.iloc[0]
    wb5.columns
    wb5=wb5.reindex(wb5.index.drop(0)).reset_index(drop=True)
    LTE_KPI=wb5.iloc[:,1:]


    wb6.columns=wb6.iloc[0]
    wb6.columns
    wb6=wb6.reindex(wb6.index.drop(0)).reset_index(drop=True)
    LTE_HO=wb6.iloc[:,1:]

    wb7.columns=wb7.iloc[0]
    wb7.columns
    wb7=wb7.reindex(wb7.index.drop(0)).reset_index(drop=True)
    LTE_RSSI=wb7.iloc[:,1:]


    WCDMA1=WCDMA[['Date','Hour','UCell Name','RRC Success Rate CS (%)','RRC Success Rate PS (%)','RAB Success Rate Speech (%)','RAB Success Rate HS (%)',            'Speech Drop Rate (%)','PS Drop Rate (%)','HS CC (%)','Soft Handover (%)','Soft HO Overhead (%)','IUCS SR (%)','IUPS SR (%)','Data Volume (GB)','Voice Traffic Erlang']]
    WCDMA_RSSI=WCDMA[['Date','Hour','UCell Name','UL RSSI (dBm)']]

    LTE_RSSI=LTE_RSSI[['Date','Hour','EUtranCell Id','UL RSSI']]

    WCDMA_peak=wb9.iloc[2:,1:]
    WCDMA_peak.columns=WCDMA_peak.iloc[0]
    WCDMA_peak.columns
    WCDMA_peak=WCDMA_peak.reindex(WCDMA_peak.index.drop(2)).reset_index(drop=True)

    LTE_peak=wb8
    LTE_peak.columns=wb8.iloc[0]
    LTE_peak.columns
    LTE_peak=LTE_peak.reindex(LTE_peak.index.drop(0)).reset_index(drop=True)

    CA=wb10.iloc[0:]
    CA.columns=CA.iloc[0]
#CA.columns
    CA=CA.reindex(CA.index.drop(0)).reset_index(drop=True)

    GSM_ngb=wb11g.iloc[2:,1:]
    GSM_ngb.columns=GSM_ngb.iloc[0]
    GSM_ngb=GSM_ngb.reindex(GSM_ngb.index.drop(2)).reset_index(drop=True)

    LTE_ngb=wb12.iloc[0:,1:]
#LTE_ngb
    LTE_ngb.columns=LTE_ngb.iloc[0]
#LTE_ngb.columns
    LTE_ngb=LTE_ngb.reindex(LTE_ngb.index.drop(0)).reset_index(drop=True)


# In[11]:


    siteid=pd.read_excel(INPATH+"siteid.xlsx",sheet_name="Sheet1")
    for sitename in siteid["Sitename"]:
    
    
        sec1_4g_1=pd.DataFrame([])
        sec2_4g_1=pd.DataFrame([])
        sec3_4g_1=pd.DataFrame([])
        sec4_4g_1=pd.DataFrame([])
        sec5_4g_1=pd.DataFrame([])
        sec6_4g_1=pd.DataFrame([])
        PCI_T=pd.DataFrame([])
        enodeBid=pd.DataFrame([])
        LTE_TAC=pd.DataFrame([])

        sec1_3g_1=pd.DataFrame([])
        sec2_3g_1=pd.DataFrame([])
        sec3_3g_1=pd.DataFrame([])
        sec4_3g_1=pd.DataFrame([])
        sec5_3g_1=pd.DataFrame([])
        sec6_3g_1=pd.DataFrame([])
        PSC_T=pd.DataFrame([])
        cell_id_2g=pd.DataFrame([])
        UMTS_LAC=pd.DataFrame([])

        sec1_3g_cellid=pd.DataFrame([])
        sec2_3g_cellid=pd.DataFrame([])
        sec3_3g_cellid=pd.DataFrame([])
        sec4_3g_cellid=pd.DataFrame([])
        sec5_3g_cellid=pd.DataFrame([])
        sec6_3g_cellid=pd.DataFrame([])


        LATITUDE=pd.DataFrame([])
#LATITUDE=pd.DataFrame([])

        LONGITUDE=pd.DataFrame([])

        cell_name_2g=pd.DataFrame([])
        BCCH=pd.DataFrame([])

        cell_id_2g=pd.DataFrame([])
        GSM_LAC=pd.DataFrame([])

        start = datetime.now()
        date=start.strftime("%d%m%Y")
        Final_site_name="CL"+"_"+sitename+"_OQA IBS SSV Submission_V1"+"_"+date+".xlsx"
######################GSM OSS KPI#################    

        sitename_gsm=sitename.replace('S',"")
        cellname_2g=GSM[GSM["Cell Name"].str.contains(sitename_gsm)]
        Hour_wise_2g=pd.DataFrame([])
        if cellname_2g.empty==False:
#Hour_wise_2g_1=pd.DataFrame([])
            Hour_count=cellname_2g['Hour'].iloc[0]
    
            First_date=cellname_2g['Date'].iloc[0]
            Date_counter=First_date
            if Hour_count==0:
                for i in range(24):
                    Date_wise_data=cellname_2g.loc[cellname_2g['Date'] == First_date]
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_2g=Hour_wise_2g.append(Hour_wise)
            else:    
                Total_hour=24
                Date_wise_data=cellname_2g.loc[cellname_2g['Date'] == First_date]
                Hour_present_dup=Date_wise_data['Hour']
                Hour_present_redup=Hour_present_dup.drop_duplicates(keep='first')
    #Hour_present_redup_1=pd.to_numeric(Hour_present_redup)
    #print(type(Hour_present_redup))
    
                Hour_present=Hour_present_redup.values
    
                for i in Hour_present:
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_2g=Hour_wise_2g.append(Hour_wise)
    
                Reqd_hour=Total_hour-len(Hour_present)
    #Date_counter=datetime.strptime(Date_counter, "%Y/%m/%d")
                Date_counter=Date_counter+timedelta(days=1)
                i+=1
            #Date_counter=Date_counter.strftime('%Y/%m/%d')
    #print(Date_counter)
    #print(type(Date_counter))
                Date_wise_data_new=cellname_2g.loc[cellname_2g['Date'] == Date_counter]
    #print(cellname_2g)
                for i in range(Reqd_hour):
                    Hour_wise_2=Date_wise_data_new.loc[Date_wise_data_new['Hour'] == i+1]
                    Hour_wise_2g=Hour_wise_2g.append(Hour_wise_2)
   # print(Hour_wise_2g)
######################GSM I->O#####################################        
        sitename_gsm=sitename.replace('S',"")
        cellname_2g_1=GSM_HO[GSM_HO["Cell Name"].str.contains(sitename_gsm)]
        Hour_wise_2g_HO=pd.DataFrame([])
#Hour_wise_2g_1=pd.DataFrame([])
        Date_wise_data_2g_HO=pd.DataFrame([])
        if cellname_2g_1.empty==False:
        #Hour_count=cellname_2g_1['Hour'].iloc[0]
            First_date=cellname_2g_1['Date'].iloc[0]
            Date_counter=First_date
        #if Hour_count==0:
        #for i in range(24):
            Date_wise_data_2g_HO=cellname_2g_1.loc[cellname_2g_1['Date'] == First_date]
            #Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
            #Hour_wise_2g_HO=Hour_wise_2g_HO.append(Hour_wise)
   # print(Date_wise_data_2g_HO)
###################WCDMA OSS KPI####################################
        cellname_3g=WCDMA1[WCDMA1["UCell Name"].str.contains(sitename)]
        Hour_wise_3g=pd.DataFrame([])
        if cellname_3g.empty==False:
    
#Hour_wise_2g_1=pd.DataFrame([])
            Hour_count=cellname_3g['Hour'].iloc[0]
            First_date=cellname_3g['Date'].iloc[0]
            Date_counter=First_date
            if Hour_count==0:
                for i in range(24):
                    Date_wise_data=cellname_3g.loc[cellname_3g['Date'] == First_date]
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_3g=Hour_wise_3g.append(Hour_wise)
            else:    
                Total_hour=24
                Date_wise_data=cellname_3g.loc[cellname_3g['Date'] == First_date]
                Hour_present_dup=Date_wise_data['Hour']
                Hour_present_redup=Hour_present_dup.drop_duplicates(keep='first')
    #Hour_present_redup_1=pd.to_numeric(Hour_present_redup)
    #print(type(Hour_present_redup))
    
                Hour_present=Hour_present_redup.values
    
                for i in Hour_present:
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_3g=Hour_wise_3g.append(Hour_wise)
    
                Reqd_hour=Total_hour-len(Hour_present)
    #Date_counter=datetime.strptime(Date_counter, "%Y/%m/%d")
                Date_counter=Date_counter+timedelta(days=1)
                i+=1
            #Date_counter=Date_counter.strftime('%Y/%m/%d')
    #print(Date_counter)
    #print(type(Date_counter))
                Date_wise_data_new=cellname_3g.loc[cellname_3g['Date'] == Date_counter]
    #print(cellname_2g)
                for i in range(Reqd_hour):
                    Hour_wise_2=Date_wise_data_new.loc[Date_wise_data_new['Hour'] == i+1]
                    Hour_wise_3g=Hour_wise_3g.append(Hour_wise_2)
#Date_wise_data_new
#Reqd_hour
#Hour_present
#Total_hour

    #print(Hour_wise_3g)
    
######################WCDMA I->O####################
    
        cellname_3g1=WCDMA_HO[WCDMA_HO["UCell Id"].str.contains(sitename)]
        Hour_wise_3g_HO=pd.DataFrame([])
        if cellname_3g1.empty==False:
    
#Hour_wise_2g_1=pd.DataFrame([])
        #Hour_count=cellname_3g1['Hour'].iloc[0]
            First_date=cellname_3g1['Date'].iloc[0]
            Date_counter=First_date
            Date_wise_data_3g=cellname_3g1.loc[cellname_3g1['Date'] == First_date]
    #print(Date_wise_data_3g)
############### LTE OSS KPI#############################
    
        cellname_4g=LTE_KPI.loc[LTE_KPI["EUtranCellFDD"].str.contains(sitename)]
        Hour_wise_4g=pd.DataFrame([])
        if cellname_4g.empty==False:
    
            Hour_wise_4g_1=pd.DataFrame([])
            Hour_count=cellname_4g['Hour'].iloc[0]
            First_date=cellname_4g['Date'].iloc[0]
            Date_counter=First_date
            if Hour_count==0:
                for i in range(24):
                    Date_wise_data=cellname_4g.loc[cellname_4g['Date'] == First_date]
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_4g=Hour_wise_4g.append(Hour_wise)
            else:    
                Total_hour=24
                Date_wise_data=cellname_4g.loc[cellname_4g['Date'] == First_date]
                Hour_present_dup=Date_wise_data['Hour']
                Hour_present_redup=Hour_present_dup.drop_duplicates(keep='first')
    #Hour_present_redup_1=pd.to_numeric(Hour_present_redup)
    #print(type(Hour_present_redup))
    
                Hour_present=Hour_present_redup.values
    
                for i in Hour_present:
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_4g=Hour_wise_4g.append(Hour_wise)
    
                Reqd_hour=Total_hour-len(Hour_present)
    #Date_counter=datetime.strptime(Date_counter, "%Y/%m/%d")
                Date_counter=Date_counter+timedelta(days=1)
                i+=1
            #Date_counter=Date_counter.strftime('%Y/%m/%d')
    #print(Date_counter)
    #print(type(Date_counter))
                Date_wise_data_new=cellname_4g.loc[cellname_4g['Date'] == Date_counter]
    #print(cellname_2g)
                for i in range(Reqd_hour):
                    Hour_wise_2=Date_wise_data_new.loc[Date_wise_data_new['Hour'] == i]
                    Hour_wise_4g=Hour_wise_4g.append(Hour_wise_2)
    #print(Hour_wise_4g)
#####################LTE I->O##########################
    
        cellname_4g1=LTE_HO[LTE_HO["EUtranCellFDD"].str.contains(sitename)]
        Hour_wise_4g_HO=pd.DataFrame([])
        if cellname_4g1.empty==False:
    
#Hour_wise_4g_1=pd.DataFrame([])
        #Hour_count=cellname_4g['Hour'].iloc[0]
            First_date=cellname_4g1['Date'].iloc[0]
            Date_counter=First_date
        
            Date_wise_data_4g=cellname_4g1.loc[cellname_4g1['Date'] == First_date]
    #print(Date_wise_data_4g)

#######################LTE RSSI#########################
        
        cellname_4g2=LTE_RSSI[LTE_RSSI["EUtranCell Id"].str.contains(sitename)]
        Hour_wise_4g_RSSI=pd.DataFrame([])
        if cellname_4g2.empty==False:
    
#Hour_wise_4g_1=pd.DataFrame([])
            Hour_count=cellname_4g2['Hour'].iloc[0]
            First_date=cellname_4g2['Date'].iloc[0]
            Date_counter=First_date
            if Hour_count==0:
                for i in range(24):
                    Date_wise_data=cellname_4g2.loc[cellname_4g2['Date'] == First_date]
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_4g_RSSI=Hour_wise_4g_RSSI.append(Hour_wise)
            else:    
                Total_hour=24
                Date_wise_data=cellname_4g2.loc[cellname_4g2['Date'] == First_date]
                Hour_present_dup=Date_wise_data['Hour']
                Hour_present_redup=Hour_present_dup.drop_duplicates(keep='first')
    #Hour_present_redup_1=pd.to_numeric(Hour_present_redup)
    #print(type(Hour_present_redup))
    
                Hour_present=Hour_present_redup.values
    
                for i in Hour_present:
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_4g_RSSI=Hour_wise_4g_RSSI.append(Hour_wise)
    
                Reqd_hour=Total_hour-len(Hour_present)
    #Date_counter=datetime.strptime(Date_counter, "%Y/%m/%d")
                Date_counter=Date_counter+timedelta(days=1)
                i+=1
            #Date_counter=Date_counter.strftime('%Y/%m/%d')
    #printnt(Date_counter)
    #print(type(Date_counter))
                Date_wise_data_new=cellname_4g2.loc[cellname_4g2['Date'] == Date_counter]
    #print(cellname_2g)
                for i in range(Reqd_hour):
                    Hour_wise_2=Date_wise_data_new.loc[Date_wise_data_new['Hour'] == i]
                    Hour_wise_4g_RSSI=Hour_wise_4g_RSSI.append(Hour_wise_2)
   # print( Hour_wise_4g_RSSI)
    
#################WCDMA RSSI###########################
        cellname_3g=WCDMA_RSSI[WCDMA_RSSI["UCell Name"].str.contains(sitename)]
        Hour_wise_3g_RSSI=pd.DataFrame([])
        if cellname_3g.empty==False:
    
#Hour_wise_4g_1=pd.DataFrame([])
            Hour_count=cellname_3g['Hour'].iloc[0]
            First_date=cellname_3g['Date'].iloc[0]
            Date_counter=First_date
            if Hour_count==0:
                for i in range(24):
                    Date_wise_data=cellname_3g.loc[cellname_3g['Date'] == First_date]
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_3g_RSSI=Hour_wise_3g_RSSI.append(Hour_wise)
            else:    
                Total_hour=24
                Date_wise_data=cellname_3g.loc[cellname_3g['Date'] == First_date]
                Hour_present_dup=Date_wise_data['Hour']
                Hour_present_redup=Hour_present_dup.drop_duplicates(keep='first')
    #Hour_present_redup_1=pd.to_numeric(Hour_present_redup)
    #print(type(Hour_present_redup))
    
                Hour_present=Hour_present_redup.values
    
                for i in Hour_present:
                    Hour_wise=Date_wise_data.loc[Date_wise_data['Hour'] == i]
                    Hour_wise_3g_RSSI=Hour_wise_3g_RSSI.append(Hour_wise)
        
                Reqd_hour=Total_hour-len(Hour_present)
    #Date_counter=datetime.strptime(Date_counter, "%Y/%m/%d")
                Date_counter=Date_counter+timedelta(days=1)
                i+=1
            #Date_counter=Date_counter.strftime('%Y/%m/%d')
    #print(Date_counter)
    #print(type(Date_counter))
                Date_wise_data_new=cellname_3g.loc[cellname_3g['Date'] == Date_counter]
    #print(cellname_2g)
                for i in range(Reqd_hour):
                    Hour_wise_2=Date_wise_data_new.loc[Date_wise_data_new['Hour'] == i]
                    Hour_wise_3g_RSSI=Hour_wise_3g_RSSI.append(Hour_wise_2)
    #print(Hour_wise_3g_RSSI)
    
                
############################LTE PEAK#################################

#LTE_peak_mod=LTE_peak.iloc[:,:]
        cellname_4g=LTE_peak[LTE_peak["ERBS Id"].str.contains(sitename)]
        LTE=pd.DataFrame([])
        if cellname_4g.empty==False:

            cellname_4g_1=cellname_4g

            cellname_4g_123=cellname_4g_1.drop_duplicates(subset="EUtranCell Id",keep='first')

            cell_details_4g=cellname_4g_123['EUtranCell Id']

            cell_details_4g=cell_details_4g.tolist()

            Date_wise_4G_DL_value=pd.DataFrame([])
            Date_wise_4G_UL_value=pd.DataFrame([])
            First_date=cellname_4g_1['Date'].iloc[0]
            Date_counter=First_date
            for i in range(4):
                Date_wise_4G=cellname_4g_1.loc[cellname_4g_1['Date'] == Date_counter]
    

                for j in range(len(cell_details_4g)):    
                    Date_wise_4G_DL_1=Date_wise_4G[Date_wise_4G["EUtranCell Id"].str.contains(cell_details_4g[j])]
        
                    Date_wise_4G_sort_Peak_DL=Date_wise_4G_DL_1.sort_values(by=["Peak DL Throughput (Mbps)"],ascending=False)
                    Date_wise_4G_DL=Date_wise_4G_sort_Peak_DL[['Date','EUtranCell Id','Peak DL Throughput (Mbps)']].iloc[0]
                    Date_wise_4G_DL_value=Date_wise_4G_DL_value.append(Date_wise_4G_DL)
    
                    Date_wise_4G_sort_Peak_UL=Date_wise_4G_DL_1.sort_values(by=["Peak UL Throughput (Mbps)"],ascending=False)
                    Date_wise_4G_UL=Date_wise_4G_sort_Peak_UL[['Date','EUtranCell Id','Peak UL Throughput (Mbps)']].iloc[0]
                    Date_wise_4G_UL_value=Date_wise_4G_UL_value.append(Date_wise_4G_UL)
    
                Date_counter=Date_counter+timedelta(days=1)
    
        


            LTE=pd.merge(Date_wise_4G_DL_value,Date_wise_4G_UL_value,on=['Date','EUtranCell Id'])
    #print(LTE)
    
 #################WCDMA_peak###################################

        cellname_3g_12=WCDMA_peak[WCDMA_peak["Ucell Name"].str.contains(sitename,na=False)]
        if cellname_3g_12.empty==False:
    #cellname_3g_12
        #cellname_3g_12=cellname_3g_12.replace("#DIV/0",0)

            cellname_3g_1=cellname_3g_12.drop_duplicates(subset="Ucell Name",keep='first')
            cell_details=cellname_3g_1['Ucell Name']
            cell_details=cell_details.tolist()

            Date_wise_3G_HS_cell_value=pd.DataFrame([])
            Date_wise_3G_EUL_user_value=pd.DataFrame([])
            Date_wise_3G_EUL_cell_value=pd.DataFrame([])
            Date_wise_3G_HS_user_value=pd.DataFrame([])
            Date_wise_3G_UL_value=pd.DataFrame([])
            First_date=cellname_3g_12['Date'].iloc[0]
            Date_counter=First_date
            for i in range(4):
                Date_wise_3G=cellname_3g_12.loc[cellname_3g_12['Date'] == Date_counter]
    

                for j in range(len(cell_details)):    
                    Date_wise_3G_1=Date_wise_3G[Date_wise_3G["Ucell Name"].str.contains(cell_details[j])]
        
                    Date_wise_3G_HS_user=Date_wise_3G_1.sort_values(by=["HS User Throughput (kbps)"],ascending=False)
                    Date_wise_3G_HS_user_1=Date_wise_3G_HS_user[['Date','Ucell Name','HS User Throughput (kbps)']].iloc[0]
                    Date_wise_3G_HS_user_value=Date_wise_3G_HS_user_value.append(Date_wise_3G_HS_user_1)
                    Date_wise_3G_HS_user_value_1=Date_wise_3G_HS_user_value[['Date','Ucell Name','HS User Throughput (kbps)']]
        
                    Date_wise_3G_HS_cell=Date_wise_3G_1.sort_values(by=["HS Cell Throughput (kbps)"],ascending=False)
                    Date_wise_3G_HS_cell_1=Date_wise_3G_HS_cell[['Date','Ucell Name','HS Cell Throughput (kbps)']].iloc[0]
                    Date_wise_3G_HS_cell_value=Date_wise_3G_HS_cell_value.append(Date_wise_3G_HS_cell_1)
                    Date_wise_3G_HS_cell_value_1=Date_wise_3G_HS_cell_value[['Date','Ucell Name','HS Cell Throughput (kbps)']]
        
                    Date_wise_3G_EUL_cell=Date_wise_3G_1.sort_values(by=["EUL Cell Throughput (kbps)"],ascending=False)
                    Date_wise_3G_EUL_cell_1=Date_wise_3G_EUL_cell[['Date','Ucell Name','EUL Cell Throughput (kbps)']].iloc[0]
                    Date_wise_3G_EUL_cell_value=Date_wise_3G_EUL_cell_value.append(Date_wise_3G_EUL_cell_1)
                    Date_wise_3G_EUL_cell_value_1=Date_wise_3G_EUL_cell_value[['Date','Ucell Name','EUL Cell Throughput (kbps)']]
        
                    Date_wise_3G_EUL_user=Date_wise_3G_1.sort_values(by=["EUL User Throughput (kbps)"],ascending=False)
                    Date_wise_3G_EUL_user_1=Date_wise_3G_EUL_user[['Date','Ucell Name','EUL User Throughput (kbps)']].iloc[0]
                    Date_wise_3G_EUL_user_value=Date_wise_3G_EUL_user_value.append(Date_wise_3G_EUL_user_1)
                    Date_wise_3G_EUL_user_value_1=Date_wise_3G_EUL_user_value[['Date','Ucell Name','EUL User Throughput (kbps)']]
        
                Date_counter=Date_counter+timedelta(days=1)
        
    #i+=1
    #Date_counter=Date_counter.strftime('%Y/%m/%d')
#type(Date_wise_3G_EUL_user_value_1)

        a=pd.merge(Date_wise_3G_HS_user_value_1,Date_wise_3G_HS_cell_value_1,on=['Date','Ucell Name'])
        b=pd.merge(Date_wise_3G_EUL_cell_value_1,Date_wise_3G_EUL_user_value_1,on=['Date','Ucell Name'])

        WCDMA=pd.merge(a,b,on=['Date','Ucell Name'])

   # print(WCDMA)

###########################CA######################

        CA_cell_wise=pd.DataFrame([])
        CA_cell=CA[CA["EUtranCell Id"].str.contains(sitename)]
        if CA_cell.empty==False:
            First_date=CA_cell['Date'].iloc[0]
            Date_wise_data=CA_cell.loc[CA_cell['Date'] == First_date]
            CA_cell_wise=CA_cell_wise.append(Date_wise_data)
   # print(CA_cell_wise)
    
############GSM 0->I###########################
        GSM_cell_wise=pd.DataFrame([])
        sitename_gsm=sitename.replace('S',"")
        GSM_cell=GSM_ngb[GSM_ngb["Adjacent Cell Name"].str.contains(sitename_gsm)]
        if GSM_cell.empty==False:
            First_date=GSM_cell['Date'].iloc[0]
            Date_wise_data=GSM_cell.loc[GSM_cell['Date'] == First_date]
            GSM_cell_wise=GSM_cell_wise.append(Date_wise_data)
    #print(GSM_cell_wise)
    
############SRS###########################
    
        wb=pd.read_excel(INPATH+"SRS_DATABASE.xlsx",sheet_name="Outdoor Site Data")
        wb.columns=wb.iloc[0]
        #wb.columns
        wb=wb.reindex(wb.index.drop(0)).reset_index(drop=True)
        wb1=wb.iloc[:,1:]
        site_name=wb1.loc[wb1['SITE ID']==sitename]
        if site_name.empty==False:
#site_name_transposed=site_name_2G.T
            LATITUDE=site_name['LATITUDE']
            LATITUDE=LATITUDE.drop_duplicates(keep = 'first') 
#LATITUDE=LATITUDE.round(3)

            LONGITUDE=site_name['LONGITUDE']
            LONGITUDE=LONGITUDE.drop_duplicates(keep = 'first') 
#LONGITUDE=LONGITUDE.round(3)

#AZIMUTH=site_name['AZIMUTH']
#AZIMUTH=AZIMUTH.drop_duplicates(keep = 'first') 


            site_name_2G=site_name.loc[site_name["TECHNOLOGY"] == 'GSM']
            site_name_transposed=site_name_2G.T
            cell_name_2g=pd.DataFrame(site_name_transposed.loc['CELL OR SECTOR ID'])
            BCCH=pd.DataFrame(site_name_transposed.loc['FREQUENCY OR FREQUENCY CHANNEL'])
            cell_id_2g=pd.DataFrame(site_name_transposed.loc['CELL NUMBER OR CID'])
            BCCH=BCCH.T
            cell_name_2g=cell_name_2g.T
            cell_id_2g=cell_id_2g.T
#LATITUDE=LATITUDE.T
#LONGITUDE=LONGITUDE.T
#AZIMUTH_T=AZIMUTH.T
            GSM_LAC=site_name_2G['LAC']
            GSM_LAC=GSM_LAC.drop_duplicates(keep = 'first') 

            site_name_3G=site_name.loc[site_name['TECHNOLOGY'] == 'WCDMA']
#site_name_3G=Tech_wise_data.loc[Tech_wise_data['SITE ID']==sitename]
#copying the 3G cell_name in 4 different df.
#list=['_01','_02','_03','_04']
#sector_3g_cell=pd.DataFrame([])
            UMTS_LAC=site_name_3G['LAC']
            UMTS_LAC=UMTS_LAC.drop_duplicates(keep = 'first') 


            sec1_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_01')]
            sec2_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_02')]
            sec3_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_03')]
            sec4_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_04')]
            sec5_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_05')]
            sec6_3g=site_name_3G[site_name_3G['CELL OR SECTOR ID'].str.contains('_06')]

            sec1_3g_1=sec1_3g["CELL OR SECTOR ID"]
            sec2_3g_1=sec2_3g['CELL OR SECTOR ID']
            sec3_3g_1=sec3_3g['CELL OR SECTOR ID']
            sec4_3g_1=sec4_3g['CELL OR SECTOR ID']
            sec5_3g_1=sec5_3g['CELL OR SECTOR ID']
            sec6_3g_1=sec6_3g['CELL OR SECTOR ID']
    
            sec1_3g_cellid=sec1_3g['CELL NUMBER OR CID']
            sec2_3g_cellid=sec2_3g['CELL NUMBER OR CID']
            sec3_3g_cellid=sec3_3g['CELL NUMBER OR CID']
            sec4_3g_cellid=sec4_3g['CELL NUMBER OR CID']
            sec5_3g_cellid=sec5_3g['CELL NUMBER OR CID']
            sec6_3g_cellid=sec6_3g['CELL NUMBER OR CID']


            PSC=site_name_3G['PSC_PCI']
            PSC=PSC.drop_duplicates(keep = 'first') 
            PSC=pd.DataFrame(PSC.astype(int))
            PSC_T=PSC.T

            site_name_4G=site_name.loc[site_name["TECHNOLOGY"] == 'LTE']
#site_name_4G=Tech_wise_data.loc[Tech_wise_data['SITE ID']==sitename]
#site_name_4G
#list=['_01','_02','_03','_04']
            sec1_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_01')]
            sec2_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_02')]
            sec3_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_03')]
            sec4_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_04')]
            sec5_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_05')]
            sec6_4g=site_name_4G[site_name_4G['CELL OR SECTOR ID'].str.contains('_06')]
#sec3_4g

            LTE_TAC=site_name_4G['LAC']
            LTE_TAC=LTE_TAC.drop_duplicates(keep = 'first') 
                        
            PCI=site_name_4G['PSC_PCI']
            PCI=PCI.drop_duplicates(keep = 'first') 
            PCI=pd.DataFrame(PCI.astype(int))
            PCI_T=PCI.T
#PCI_T


            enodeBid=pd.DataFrame(site_name_4G['eNodeB ID'])
            enodeBid=enodeBid.drop_duplicates(keep = 'first') 
            enodeBid=enodeBid.astype(int)
            enodeBid=enodeBid.T
            enodeb=enodeBid.to_string()
#type(enodeBid)
    
    
            sec1_4g_1=sec1_4g["CELL OR SECTOR ID"]
            sec2_4g_1=sec2_4g['CELL OR SECTOR ID']
            sec3_4g_1=sec3_4g['CELL OR SECTOR ID']
            sec4_4g_1=sec4_4g['CELL OR SECTOR ID']
            sec5_4g_1=sec5_4g['CELL OR SECTOR ID']
            sec6_4g_1=sec6_4g['CELL OR SECTOR ID']
    

#LTE_ngb
######################LTE O->I#################################
        LTE_cell_wise=pd.DataFrame([])
        enodeb=enodeBid.values.tolist()
        for i in enodeb:
            LTE_cell=LTE_ngb[LTE_ngb["EUtranCellRelation"].str.contains(str(i))]
            LTE_cell_wise=pd.DataFrame([])
    
            if LTE_cell.empty==False:
                First_date=LTE_cell['Date'].iloc[0]
                Date_wise_data=LTE_cell.loc[LTE_cell['Date'] == First_date]
                LTE_cell_wise=LTE_cell_wise.append(Date_wise_data)
   # print("Check the output:########",LTE_cell_wise)

################wcdma O->I#########################################
        WCDMA_1=pd.DataFrame([])
        wcdma_ngb=pd.concat([wb11,wb21,wb31,wb41])
#wcdma_ngb

        wcdma_ngb_1=wcdma_ngb.iloc[2:,1:]

        wcdma_ngb_1=wcdma_ngb_1.replace("#DIV/0","")
        wcdma_ngb_1 = wcdma_ngb_1.reset_index(drop=True)
        wcdma_ngb_1.columns=wcdma_ngb_1.iloc[0]

        wcdma_ngb_1=wcdma_ngb_1.reindex(wcdma_ngb_1.index.drop(0)).reset_index(drop=True)
#wcdma_ngb_1
        wcdma_ngb_2=wcdma_ngb_1.dropna()

    #WCDMA_1=pd.DataFrame([])
        First_date=wcdma_ngb_2['Date'].iloc[0]
        Date_wise_data=wcdma_ngb_2.loc[wcdma_ngb_2['Date'] == First_date]
        for i in site_name_3G['CELL NUMBER OR CID']:
            WCDMA_1a=wcdma_ngb_2[wcdma_ngb_2["UtranRelation"].str.endswith(i)]
            WCDMA_1=WCDMA_1.append(WCDMA_1a)
    #print(WCDMA_1)
########################################VSWR###################

        import shutil
        shutil.copy(INPATH+"IBS.xlsx",OUTPATH+"IBS_NEW1.xlsx")
        if os.path.isfile(OUTPATH+'IBS_NEW1.xlsx'):
            os.rename(OUTPATH+'IBS_NEW1.xlsx',OUTPATH+Final_site_name)


        invxrg=[]
        invxrg1=[]
        with open (INPATH+'Combined file.log', 'rt') as myfile:
            myline=myfile.readlines()
            for i in range(0,len(myline)):
                line = myline[i]
            #print(line) 
                if sitename+"_BB1> invxrg" in line:
                    for k in range(i,i+300):
                        stop_line=myline[k]
                        if sitename+"_BB1> get eutrancellfdd=* cellid" in stop_line:
                            break
                        else:
                            invxrg.append(stop_line)
                if sitename+"_BB2> invxrg" in line:
                    for k in range(i,i+300):
                        stop_line=myline[k]
                        if sitename+"_BB2> get eutrancellfdd=* cellid" in stop_line:
                            break
                        else:
                            invxrg1.append(stop_line)



        with open(OUTPATH+sitename+"BB1"+".txt", 'a') as site:
            for listitem in invxrg:
                site.write('%s\n' % listitem)
        with open(OUTPATH+sitename+"BB2"+".txt", 'a') as site:
            for listitem in invxrg1:
                site.write('%s\n' % listitem) 

        vswr=[]
        vswr1=[]

        with open (OUTPATH+sitename+"BB1"+".txt", 'rt') as myfile:
            myline=myfile.readlines()
            for i in range(0,len(myline)):
                line = myline[i]
                if "VSWR (RL)   RX (dBm)" in line:
                    for k in range(i-2,i+100):
                        stop_line=myline[k]
                    #print(stop_line) 
                        if "Tip: use option "in stop_line:
                            break
                        else:
                            vswr.append(stop_line)   
                        
        with open (OUTPATH+sitename+"BB2"+".txt", 'rt') as myfile:
            myline=myfile.readlines()
            for i in range(0,len(myline)):
                line = myline[i]
                if "VSWR (RL)   RX (dBm)" in line:
                    for k in range(i-2,i+100):
                        stop_line=myline[k]
                    #print(stop_line) 
                        if "Tip: use option "in stop_line:
                            break
                        else:
                            vswr1.append(stop_line)   
    


   

        import xlwt

        if os.path.isfile(OUTPATH+'snap.xls'):
            os.remove('snap.xls')

        wb = xlwt.Workbook()
        ws1 = wb.add_sheet('Sheet1')
        ws2 = wb.add_sheet('Sheet2')
        ws3 = wb.add_sheet('Sheet3')

        first_column = 0

# write each item in the list to consecutive columns on the first row

        for index, item in enumerate(vswr,1):
            ws1.write(index,first_column,item) 


        for index, item in enumerate(vswr1,1):
            ws2.write(index,first_column,item)

    #for index, item in enumerate(vswr2,1):
       #ws3.write(index,first_column,item) 
             
             
        wb.save(OUTPATH+'snap.xls')
    
        excel2img.export_img(OUTPATH+"snap.xls",OUTPATH+"VSWR_BB1.png","","Sheet1!A2:T50")
        excel2img.export_img(OUTPATH+"snap.xls",OUTPATH+"VSWR_BB2.png","","Sheet2!A2:T50")

        book=openpyxl.load_workbook(OUTPATH+Final_site_name)
#sheet = book.get_sheet_by_name("Site Information")

        writer=pd.ExcelWriter(OUTPATH+Final_site_name,engine='openpyxl')
        writer.book=book
        writer.sheets=dict((ws.title,ws) for ws in book.worksheets)
#print(writer.sheets)
#4G
#sitename1.to_excel(writer,"Site Information",startcol=1,startrow=0,header=None,index=False)
        CA_cell_wise[CA_cell_wise.columns[1:]].to_excel(writer,"CA",startcol=0,startrow=1,header=None,index=False)
        LTE.to_excel(writer,"LTE Peak Throughput",startcol=0,startrow=1,header=None,index=False)
        WCDMA.to_excel(writer,"WCDMA Peak Throughput",startcol=0,startrow=1,header=None,index=False)
        GSM_cell_wise.to_excel(writer,"GSM O->I",startcol=0,startrow=1,header=None,index=False)
        LTE_cell_wise.to_excel(writer,"LTE O->I",startcol=0,startrow=1,header=None,index=False)
        WCDMA_1.to_excel(writer,"WCDMA O->I",startcol=0,startrow=0,index=False)
        Hour_wise_2g.to_excel(writer,"GSM OSS KPIs",startcol=0,startrow=1,header=None,index=False)
        Date_wise_data_2g_HO.to_excel(writer,"GSM I->O",startcol=0,startrow=1,header=None,index=False)
        Hour_wise_3g.to_excel(writer,"WCDMA OSS KPIs",startcol=0,startrow=1,header=None,index=False)
        Date_wise_data_3g.to_excel(writer,"WCDMA I->O",startcol=0,startrow=1,header=None,index=False)
        Hour_wise_4g.to_excel(writer,"LTE OSS KPIs",startcol=0,startrow=1,header=None,index=False)
        Date_wise_data_4g.to_excel(writer,"LTE I->O",startcol=0,startrow=1,header=None,index=False)
        Hour_wise_4g_RSSI.to_excel(writer,"LTE UL RSSI",startcol=0,startrow=1,header=None,index=False)
        Hour_wise_3g_RSSI.to_excel(writer,"WCDMA UL RSSI",startcol=0,startrow=1,header=None,index=False)
        sec1_4g_1.to_excel(writer,"Site Information",startcol=1,startrow=17,header=None,index=False)
        sec2_4g_1.to_excel(writer,"Site Information",startcol=2,startrow=17,header=None,index=False)
        sec3_4g_1.to_excel(writer,"Site Information",startcol=3,startrow=17,header=None,index=False)
        sec4_4g_1.to_excel(writer,"Site Information",startcol=4,startrow=17,header=None,index=False)
        sec5_4g_1.to_excel(writer,"Site Information",startcol=5,startrow=17,header=None,index=False)
        sec6_4g_1.to_excel(writer,"Site Information",startcol=6,startrow=17,header=None,index=False)
        PCI_T.to_excel(writer,"Site Information",startcol=1,startrow=22,header=None,index=False)
        enodeBid.to_excel(writer,"Site Information",startcol=1,startrow=21,header=None,index=False)
        LTE_TAC.to_excel(writer,"Site Information",startcol=1,startrow=23,header=None,index=False)
#3G
        sec1_3g_1.to_excel(writer,"Site Information",startcol=1,startrow=9,header=None,index=False)
        sec2_3g_1.to_excel(writer,"Site Information",startcol=2,startrow=9,header=None,index=False)
        sec3_3g_1.to_excel(writer,"Site Information",startcol=3,startrow=9,header=None,index=False)
        sec4_3g_1.to_excel(writer,"Site Information",startcol=4,startrow=9,header=None,index=False)
        sec5_3g_1.to_excel(writer,"Site Information",startcol=5,startrow=9,header=None,index=False)
        sec6_3g_1.to_excel(writer,"Site Information",startcol=6,startrow=9,header=None,index=False)
        PSC_T.to_excel(writer,"Site Information",startcol=1,startrow=13,header=None,index=False)
        cell_id_2g.to_excel(writer,"Site Information",startcol=1,startrow=7,header=None,index=False)
        UMTS_LAC.to_excel(writer,"Site Information",startcol=1,startrow=16,header=None,index=False)

        sec1_3g_cellid.to_excel(writer,"Site Information",startcol=1,startrow=14,header=None,index=False)
        sec2_3g_cellid.to_excel(writer,"Site Information",startcol=2,startrow=14,header=None,index=False)
        sec3_3g_cellid.to_excel(writer,"Site Information",startcol=3,startrow=14,header=None,index=False)
        sec4_3g_cellid.to_excel(writer,"Site Information",startcol=4,startrow=14,header=None,index=False)
        sec5_3g_cellid.to_excel(writer,"Site Information",startcol=5,startrow=14,header=None,index=False)
        sec6_3g_cellid.to_excel(writer,"Site Information",startcol=6,startrow=14,header=None,index=False)
#2G

        LATITUDE.to_excel(writer,"Site Information",startcol=1,startrow=4,header=None,index=False)
        LATITUDE.to_excel(writer,"Site Information",startcol=2,startrow=4,header=None,index=False)
        LATITUDE.to_excel(writer,"Site Information",startcol=3,startrow=4,header=None,index=False)
        LATITUDE.to_excel(writer,"Site Information",startcol=4,startrow=4,header=None,index=False)
        LATITUDE.to_excel(writer,"Site Information",startcol=5,startrow=4,header=None,index=False)
        LATITUDE.to_excel(writer,"Site Information",startcol=6,startrow=4,header=None,index=False)

        LONGITUDE.to_excel(writer,"Site Information",startcol=1,startrow=3,header=None,index=False)
        LONGITUDE.to_excel(writer,"Site Information",startcol=2,startrow=3,header=None,index=False)
        LONGITUDE.to_excel(writer,"Site Information",startcol=3,startrow=3,header=None,index=False)
        LONGITUDE.to_excel(writer,"Site Information",startcol=4,startrow=3,header=None,index=False)
        LONGITUDE.to_excel(writer,"Site Information",startcol=5,startrow=3,header=None,index=False)
        LONGITUDE.to_excel(writer,"Site Information",startcol=6,startrow=3,header=None,index=False)


        cell_name_2g.to_excel(writer,"Site Information",startcol=1,startrow=5,header=None,index=False)
        BCCH.to_excel(writer,"Site Information",startcol=1,startrow=6,header=None,index=False)

        cell_id_2g.to_excel(writer,"Site Information",startcol=1,startrow=7,header=None,index=False)
        GSM_LAC.to_excel(writer,"Site Information",startcol=1,startrow=8,header=None,index=False)
        writer.save() 
    
        book=openpyxl.load_workbook(OUTPATH+Final_site_name)
        sheet = book.get_sheet_by_name("VSWR")
        img1 = openpyxl.drawing.image.Image(OUTPATH+'VSWR_BB1.png')
    
        sheet.add_image(img1,'A2')

        img2 = openpyxl.drawing.image.Image(OUTPATH+'VSWR_BB2.png')
        sheet.add_image(img2,'A17')

        book.save(OUTPATH+Final_site_name)
        print("One Site completed..Next started...Enjoy!!!!!")


# In[ ]:

if __name__ == '__main__':
    import sys
    if len(sys.argv)<3:
        print("Please provide input/output path")
        sys.exit()
    INPATH=sys.argv[1]
    OUTPATH=sys.argv[2]
    actual(INPATH,OUTPATH)



